from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import date

input_file = '/Users/z/Desktop/src.xlsx'
save_file = '/Users/z/Desktop/test.xlsx'

wb_input = load_workbook(input_file)
wb_output = Workbook()

sheet_name_list = wb_input.get_sheet_names()
print(sheet_name_list)

ws_input = wb_input.get_sheet_by_name(sheet_name_list[1])
print(ws_input.title)

ModelOneColIndex = 2
ModelSecondColIndex = 3
ModelThirdColIndex = 4

DataRowBeginIndex = 3

ModelTwoSecondBeginIndex = 4

FirstModelRowIndex = []
SecondModelRowIndex = []
ThirdModelRowIndex = []

FixedCellFillStyle = PatternFill(fill_type='solid', start_color='ffc0c0c0', end_color='0c0c0cff')

def readFirstModel():
    for col in ws_input.get_squared_range(ModelOneColIndex, DataRowBeginIndex, ModelOneColIndex, ws_input.max_row):
        if col[0].value != None:
            FirstModelRowIndex.append(col[0].row)

def readSecondModel():
    for col in ws_input.get_squared_range(ModelSecondColIndex, DataRowBeginIndex, ModelSecondColIndex, ws_input.max_row):
        if col[0].value != None:
            SecondModelRowIndex.append(col[0].row)

def readThirdModel():
    for col in ws_input.get_squared_range(ModelThirdColIndex, DataRowBeginIndex, ModelThirdColIndex, ws_input.max_row):
        if col[0].value != None:
            ThirdModelRowIndex.append(col[0].row)


readFirstModel()
readSecondModel()
readThirdModel()
'''
for index in SecondModelRowIndex:
    print('second model row index: ', index)
    print(ws_input.cell(row=index, column=ModelSecondColIndex).value)

'''

def writeFirstModelFixedContent(ws):
    ws['A1'].value = '一级模块:'
    ws['A1'].fill = FixedCellFillStyle
    ws['C1'].value = '编制人:黄池莲'
    #ws['C1'].fill = FixedCellFillStyle
    ws['F1'].value = '优先级:高'
    #ws['F1'].fill = FixedCellFillStyle
    ws['A2'].value  = '功能概述:'
    ws['A2'].fill = FixedCellFillStyle

def writeFirstModelDynamicContent(ws, cell):
    ws['B1'] = cell.value
    ws['D1'] = '编制时间:' + date.today().isoformat()

def writeSecondModelFixedContent(ws, base_index):
    ws['A{}'.format(base_index)].value = '二级模块:'
    ws['A{}'.format(base_index)].fill = FixedCellFillStyle

    ws['A{}'.format(base_index + 1)].value = '测试目的:'
    ws['A{}'.format(base_index + 1)].fill = FixedCellFillStyle

    ws['D{}'.format(base_index + 1)].value = '前置条件:'
    ws['D{}'.format(base_index + 1)].fill = FixedCellFillStyle

    ws['A{}'.format(base_index + 2)].value = '用例编号:'
    ws['A{}'.format(base_index + 2)].fill = FixedCellFillStyle

    ws['A{}'.format(base_index + 3)].value = '测试项'
    ws['A{}'.format(base_index + 3)].fill = FixedCellFillStyle

    ws['B{}'.format(base_index + 3)].value = '操作步骤'
    ws['B{}'.format(base_index + 3)].fill = FixedCellFillStyle

    ws['C{}'.format(base_index + 3)].value = '测试数据'
    ws['C{}'.format(base_index + 3)].fill = FixedCellFillStyle

    ws['D{}'.format(base_index + 3)].value = '期望结果'
    ws['D{}'.format(base_index + 3)].fill = FixedCellFillStyle

    ws['F{}'.format(base_index + 3)].value = '测试结果'
    ws['F{}'.format(base_index + 3)].fill = FixedCellFillStyle

def writeSecondModelDynamicContent(ws, base_index, begin, end):
    print('begin: ', begin)
    print('end:', end)
    ws['B{}'.format(base_index)].value = ws_input.cell(coordinate='C{}'.format(begin)).value
    ws['E{}'.format(base_index + 1)].value = ws_input.cell(coordinate='F{}'.format(begin)).value
    ws['B{}'.format(base_index + 2)].value = ws_input.cell(coordinate='E{}'.format(begin)).value

    third_rows = 0
    content_row_begin = 3
    for third in ThirdModelRowIndex:
        if third >= begin and third < end:
            # counter
            third_rows = third_rows + 1

            third_begin = third
            third_end = ThirdModelRowIndex[ThirdModelRowIndex.index(third)+1]
            print('third_begin: ', third_begin)
            print('third_end:', third_end)

            test_step = []
            test_data_good = []
            test_data_bad = []
            expect_data = []

            for step in range(third_begin, third_end):
                test_step.append(ws_input['G{}'.format(step)].value)
                test_data_good.append(ws_input['H{}'.format(step)].value)
                test_data_bad.append(ws_input['I{}'.format(step)].value)
                expect_data.append(ws_input['J{}'.format(step)].value)

            print('test_step:', test_step)
            print('good_data:', test_data_good)
            print('bad_data:', test_data_bad)
            print('expect_data:', expect_data)

            for i in range(0, len(test_data_good)):
                if test_data_good[i] == '/' or test_data_good[i] == None:
                    test_data_good[i] = ' '
                if test_data_bad[i] == '/' or test_data_bad[i] == None:
                    test_data_bad[i] = ' '

            for i in range(0, len(test_step)):
                if test_step[i] is None:
                    test_step[i] = ' '

            for i in range(0, len(expect_data)):
                if expect_data[i] is None:
                    expect_data[i] = ' '

            if test_step is not None:
                ws['B{}'.format(base_index + content_row_begin + third_rows)].value = '\n'.join(test_step)
            if test_data_good is not None:
                ws['C{}'.format(base_index + content_row_begin + third_rows)].value = '\n'.join(test_data_good)
            if expect_data is not None:
                ws['D{}'.format(base_index + content_row_begin + third_rows)].value = '\n'.join(expect_data)

            ws['A{}'.format(base_index + content_row_begin + third_rows)].value = ws_input.cell(coordinate='D{}'.format(third_begin)).value
    # begin row, fixed row, dynamic row, empty row, next row
    return base_index + 3 + third_rows + 1 + 1

def writeSecondModel(ws, cell, first_model_index):
    print('first model index: ', first_model_index)
    next_index = FirstModelRowIndex.index(first_model_index) + 1
    print('next index: ', next_index)
    if next_index >= len(FirstModelRowIndex):
        return

    next_model_index = FirstModelRowIndex[next_index]
    print('next model index: ', next_model_index)

    base_index = ModelTwoSecondBeginIndex
    for second in SecondModelRowIndex:
        #print('second row index', second)
        if second >= first_model_index and second < next_model_index:
            print('second row index', second)
            writeSecondModelFixedContent(ws, base_index)
            base_index = writeSecondModelDynamicContent(ws,
                                           base_index,
                                           second,
                                           SecondModelRowIndex[SecondModelRowIndex.index(second)+1])


def writeFirstModel():
    for index in FirstModelRowIndex:
        cell = ws_input.cell(row=index, column=ModelOneColIndex)
        if cell.value != None:
            print('write cell: ', cell.value)
            wb_output.create_sheet(cell.value)
            fm_ws = wb_output.get_sheet_by_name(cell.value)
            # fixed content
            writeFirstModelFixedContent(fm_ws)
            # write dynamic content
            writeFirstModelDynamicContent(fm_ws, cell)
            # process second model
            writeSecondModel(fm_ws, cell, index)


writeFirstModel()

wb_output.save(save_file)

